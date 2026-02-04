from inspect_ai.dataset import MemoryDataset, Sample
import pandas as pd
import os
from openpyxl.styles import Alignment
from data_utils.load_mcqa_task import _save_dataset
from data_utils.refine_dataset import MCQ
from utils.enums import Metrics


def create_sample_to_score_from_refined_logs(refined_report_card_logs, base_config):
    """
    Create sample_to_score dictionary from refined report_card_logs.
    This extracts ALL metric scores from the refined dataset evaluation.
    Maps eval_sample.scores directly since it already has the structure we need.
    """
    sample_to_score = {}
    
    # Extract all metric scores from refined_report_card_logs
    for eval_log in refined_report_card_logs:
        for eval_sample in eval_log.samples:
            sample_id = eval_sample.id
            
            # Map scores directly - eval_sample.scores already has the right structure
            sample_to_score[sample_id] = {'accuracy': {}}
            
            for score_name, score_value in eval_sample.scores.items():
                # Map score names to what scorers expect
                if score_name == 'diff':
                    sample_to_score[sample_id]['difficulty'] = score_value.value
                elif score_name == 'disc':
                    sample_to_score[sample_id]['discriminability'] = score_value.value
                elif score_name == 'avg_accuracy':
                    sample_to_score[sample_id]['accuracy']['avg'] = {
                        'score': score_value.value,
                        'answer': score_value.answer,
                        'explanation': score_value.explanation
                    }
                else:
                    # For other scorers, store the full structure
                    sample_to_score[sample_id][score_name] = {
                        'value': score_value.value,
                        'answer': score_value.answer,
                        'explanation': score_value.explanation,
                        'metadata': score_value.metadata
                    }
    
    return sample_to_score


def create_sample_to_score_from_report_card_logs(report_card_logs, base_config, refinement_logs=None):
    """
    Create sample_to_score dictionary from report_card_logs for original MCQs.
    This allows us to cache metric results for MCQs that weren't refined.
    Only includes cached results for MCQs that were NOT refined.
    """
    sample_to_score = {}
    
    # Create a set of sample IDs that were refined (if refinement_logs provided)
    refined_sample_ids = set()
    if refinement_logs:
        # Handle both lightweight and full refinement_logs
        samples = refinement_logs[0].samples if hasattr(refinement_logs[0], 'samples') else refinement_logs[0]
        for sample in samples:
            # Handle both lightweight and full sample objects
            if hasattr(sample, 'metadata'):
                sample_metadata = sample.metadata
                sample_id = sample.id
            else:
                # If it's already a dict-like object, use it directly
                sample_metadata = sample if isinstance(sample, dict) else {}
                sample_id = sample_metadata.get('id', None)
            
            # Check if this sample was actually refined by comparing content
            if 'question' in sample_metadata and 'old_question' in sample_metadata:
                was_refined = (sample_metadata['question'] != sample_metadata['old_question'] or 
                              sample_metadata['choices_list'] != sample_metadata['old_choices_list'] or
                              sample_metadata['target'] != sample_metadata['old_target'])
                
                if was_refined:
                    refined_sample_ids.add(sample_id)
    
    # Extract all metric scores from report_card_logs
    for eval_log in report_card_logs:
        for eval_sample in eval_log.samples:
            sample_id = eval_sample.id
            
            # Skip refined samples - don't cache their results
            if sample_id in refined_sample_ids:
                continue
                
            sample_to_score[sample_id] = {'accuracy': {}}
            
            # Extract all metric scores with their exact metadata
            for score_name, score_value in eval_sample.scores.items():
                if score_name.startswith('accuracy_'):
                    model_name = score_name[len('accuracy_'):]
                    sample_to_score[sample_id]['accuracy'][model_name] = {
                        'score': score_value.value,
                        'answer': score_value.answer,
                        'explanation': score_value.explanation
                    }
                elif score_name == 'shortcuts':
                    sample_to_score[sample_id]['shortcuts'] = {
                        'value': score_value.value,
                        'answer': score_value.answer,
                        'explanation': score_value.explanation,
                        'metadata': score_value.metadata
                    }
                elif score_name == 'contamination':
                    sample_to_score[sample_id]['contamination'] = {
                        'value': score_value.value,
                        'answer': score_value.answer,
                        'explanation': score_value.explanation,
                        'metadata': score_value.metadata
                    }
                elif score_name == 'writing_flaws':
                    sample_to_score[sample_id]['writing_flaws'] = {
                        'value': score_value.value,
                        'answer': score_value.answer,
                        'explanation': score_value.explanation,
                        'metadata': score_value.metadata
                    }
    
    # If difficulty metric is specified, we need to load IRT results for original dataset
    if Metrics.DIFFICULTY.value in base_config.get("scoring_metrics", []):
        from utils.cache import Cache, CacheType
        from model_utils.irt import PyMCIRTModel
        
        cache = Cache(cache_dir=base_config.get("cache_dir"), cache_type=CacheType(base_config.get('cache_type', 'none')))
        resolved_run_name = base_config.get("metric_run_name")
        
        # Load cached IRT results for original dataset
        cached_sample_to_score = cache.load('irt_logs', resolved_run_name, 'data_sample_to_score')
        
        if cached_sample_to_score:
            # Merge IRT results (difficulty, discriminability) with accuracy results
            # Only for non-refined samples
            for sample_id in sample_to_score:
                if sample_id in cached_sample_to_score:
                    sample_to_score[sample_id]['difficulty'] = cached_sample_to_score[sample_id].get('difficulty', 0.0)
                    sample_to_score[sample_id]['discriminability'] = cached_sample_to_score[sample_id].get('discriminability', 0.0)
    
    return sample_to_score


def merge_sample_to_score(original_sample_to_score, refined_sample_to_score, refinement_logs):
    """
    Merge original sample_to_score with refined IRT results.
    The original_sample_to_score already excludes refined MCQs, so we just add refined results.
    """
    merged_sample_to_score = original_sample_to_score.copy()
    
    # Add refined IRT results for MCQs that were actually refined
    if refined_sample_to_score:
        for sample_id, refined_data in refined_sample_to_score.items():
            merged_sample_to_score[sample_id] = refined_data
    
    return merged_sample_to_score


def save_refined_dataset(refinement_logs, base_config, refine_config, report_card_logs, should_save: bool = False):
    new_mcqs = []
    old_mcqs = []
    # Handle both lightweight and full refinement_logs
    samples = refinement_logs[0].samples if hasattr(refinement_logs[0], 'samples') else refinement_logs[0]
    for sample in samples:
        # Handle both lightweight and full sample objects
        if hasattr(sample, 'metadata'):
            sample_metadata = sample.metadata
        else:
            # If it's already a dict-like object, use it directly
            sample_metadata = sample if isinstance(sample, dict) else {}
        
        if sample_metadata.get('should_skip', False) and 'question' not in sample_metadata: # skip questions that should be filtered
            continue

        new_mcq = MCQ(question=sample_metadata['question'], choices=sample_metadata['choices_list'], answer=sample_metadata['target'])
        new_mcqs.append(new_mcq)

        old_mcq = MCQ(question=sample_metadata['old_question'], choices=sample_metadata['old_choices_list'], answer=sample_metadata['old_target'])
        old_mcqs.append(old_mcq)

    if should_save:
        saved_path = _save_dataset(
            [mcq.to_json() for mcq in new_mcqs], 
            base_config.get('dataset_save_dir'), 
            base_config.get('refine_run_name')
        )
        save_annotations_excel(old_mcqs, new_mcqs, base_config)
    
    # Create sample_to_score from report_card_logs for caching original MCQ results
    sample_to_score = create_sample_to_score_from_report_card_logs(report_card_logs, base_config, refinement_logs)
    
    return MemoryDataset([Sample(input=mcq.question, choices=mcq.choices, target=mcq.answer) for mcq in new_mcqs]), sample_to_score


def save_annotations_excel(old_mcqs, new_mcqs, base_config):
    annot_df = {'old_mcq': [], 'new_mcq': [], 'Is Valid?': []}
    for old_mcq, new_mcq in zip(old_mcqs, new_mcqs):
        annot_df['old_mcq'].append(old_mcq.to_prompt())
        annot_df['new_mcq'].append(new_mcq.to_prompt())
        annot_df['Is Valid?'].append('')
    annot_df = pd.DataFrame(annot_df)
    
    subset_df = annot_df.sample(n=min(50, len(annot_df)), random_state=42).copy()
    
    annot_path = os.path.join(base_config.get('dataset_save_dir'), base_config.get('refine_run_name'), 'annotations.xlsx')
    with pd.ExcelWriter(annot_path, engine='openpyxl') as writer:
        subset_df.to_excel(writer, index=False, sheet_name='Subset (50)')
        annot_df.to_excel(writer, index=False, sheet_name='All Annotations')
        
        subset_worksheet = writer.sheets['Subset (50)']
        all_worksheet = writer.sheets['All Annotations']
        
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        for worksheet in [subset_worksheet, all_worksheet]:
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = wrap_alignment
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 20), 80)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        last_row = len(subset_df) + 1
        
        summary_row = last_row + 2
        subset_worksheet.cell(row=summary_row, column=1, value="Validation Summary:")
        subset_worksheet.cell(row=summary_row, column=2, value="Valid Count:")
        
        valid_count_formula = f"=COUNTIF(D2:D{last_row},1)"
        subset_worksheet.cell(row=summary_row, column=3, value=valid_count_formula)
        
        subset_worksheet.cell(row=summary_row + 1, column=2, value="Valid Percentage:")
        percentage_formula = f"=C{summary_row}/{len(subset_df)}"
        subset_worksheet.cell(row=summary_row + 1, column=3, value=percentage_formula)
        
        for row in [summary_row, summary_row + 1]:
            for col in [1, 2, 3]:
                cell = subset_worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(vertical='top')
                if col == 1 or col == 2:
                    cell.font = cell.font.copy(bold=True)
    
    print("Saving data to annotate:", annot_path)
